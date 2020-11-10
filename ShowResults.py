import matplotlib.pyplot as plt
import openpyxl
from glob import glob
import numpy as np


def autolabel(rects, all_mse):
    """
    Attach a text label above each bar displaying its height
    """
    for rect, mse in zip(rects, all_mse):
        height = rect.get_height()
        ax.text(rect.get_x() + rect.get_width()/2., 1.001*height,
                "%.4f" % np.round(mse, 4),
                ha='center', va='bottom', fontsize=14)


def compute_measures(exp_number, all_features_extraction_methods, dataset_name, test_sets, sub_exp_number=None):

    if sub_exp_number is not None:
        folder_name1 = "Results/Exp" + str(exp_number) + "/" + str(sub_exp_number)
    else:
        folder_name1 = "Results/Exp" + str(exp_number)
    folder_name2 = "Descriptors/"
    image_name = folder_name1 + "/*.npz"
    selected_algorithms = glob(image_name)

    legend_titles = []
    all_mse = []
    all_r2_percent = []
    all_test_elapsed_time = []
    all_desc_elapsed_time = []
    legends = []
    for algorithmNumber in range(len(selected_algorithms)):
        file_name = selected_algorithms[algorithmNumber]
        index = [file_name.find(x) for x in all_features_extraction_methods if file_name.find(x) != -1][0]
        legend_titles.append(file_name[index:].split(".")[0])
        legends.append(file_name[file_name.find(dataset_name):].split(".")[0])
        print(legends[algorithmNumber])

        results = np.load(file_name)

        if exp_number not in (1, 4, 9, 10):
            descriptor_file = np.load(folder_name2 + dataset_name + "_" + legend_titles[algorithmNumber] + "_descriptor.npz")
            desc_elapsed_time = np.mean(descriptor_file["elapsed_time"])
            all_desc_elapsed_time.append(desc_elapsed_time)
            print(desc_elapsed_time)

        test_elapsed_time = np.mean(results["test_elapsed_time"])
        all_test_elapsed_time.append(test_elapsed_time)
        print(test_elapsed_time)
        a = results["testing_labels"]
        b = results["training_labels"]
        mse = results["squared_mean_error"]
        all_mse.append(mse)
        print(mse)

        r_score_percent = results["r_score"] * 100
        all_r2_percent.append(r_score_percent)
        print(r_score_percent)

    file_name = "_".join([folder_name1 + folder_name1.replace("/", "_"), "Results.xlsx"])

    wb = openpyxl.Workbook()

    wb.save(file_name)

    wb = openpyxl.load_workbook(filename=file_name)
    ws = wb.worksheets[0]

    ws["B" + str(1)] = "MSE"
    ws["C" + str(1)] = "Time (Feature Extraction)"
    ws["D" + str(1)] = "Time (GP Test)"
    ws["E" + str(1)] = "R %"

    for ii in range(len(all_mse)):
        ws["A" + str(ii + 2)] = legends[ii]
        ws["B" + str(ii + 2)] = str(all_mse[ii])
        ws["C" + str(ii + 2)] = str(all_test_elapsed_time[ii])

        if len(all_desc_elapsed_time) >= 1:
            ws["D" + str(ii + 2)] = str(all_desc_elapsed_time[ii])

        ws["E" + str(ii + 2)] = str(all_r2_percent[ii])

    wb.save(file_name)

    return all_mse


expNumber = 8
sub_exp_number1 = 1
sub_exp_number2 = 2
features_extraction_number = 0
dataset_name = "Noisy"
n_groups = 6

all_features_extraction_methods = ["Raw", "Hog", "ResNet18"]
features_extraction_method = all_features_extraction_methods[features_extraction_number]
all_titles = ["", "Raw Pixel Values", "Hog", "ResNet18", "Hog, ResNet18 and Raw Pixel Values",
              "Hog", "ResNet18", "Hog and ResNet18", "Raw Pixel Values", "Hog, ResNet18 and Raw Pixel Values"]
part_of_title = all_titles[expNumber-1]
test_sets = ["JPEG_Compression", "Motion_Blur", "Mild_Gaussian", "Strong_Gaussian", "Normal"]

all_mse1 = compute_measures(expNumber, all_features_extraction_methods, dataset_name, test_sets, sub_exp_number1)
all_mse2 = compute_measures(expNumber, all_features_extraction_methods, dataset_name, test_sets, sub_exp_number2)

if expNumber in (2, 5):
    all_mse1[0], all_mse1[1], all_mse1[2] = all_mse1[2], all_mse1[0], all_mse1[1]
    all_mse2[0], all_mse2[1], all_mse2[2] = all_mse2[2], all_mse2[0], all_mse2[1]

if expNumber == 8:
    all_mse1[2], all_mse1[3], all_mse1[5] = all_mse1[5], all_mse1[2], all_mse1[3]
    all_mse2[2], all_mse2[3], all_mse2[5] = all_mse2[5], all_mse2[2], all_mse2[3]

if expNumber == 10:
    all_mse2[1], all_mse2[2] = all_mse2[2], all_mse2[1]

# create plot
fig, ax = plt.subplots()

index = np.arange(n_groups)
bar_width = 0.35
opacity = 0.8
first_label = ""
second_label = ""
if expNumber == 1:
    plt.xticks(index + bar_width / 2, ("Raw"), fontsize=14)
    x_label = "Raw"

if expNumber in (2, 5):
    plt.xticks(index + bar_width / 2, ('8*8', '16*16', '32*32'), fontsize=14)
    x_label = "Cell Size"
    first_label = '9 Features Per Cell'
    second_label = '128 Features Per Cell'

elif expNumber in (3, 6):
    plt.xticks(index + bar_width, ('Block 3', 'Block 4', 'Block 5'), fontsize=14)
    x_label = "Block Number"

elif expNumber in (4, 10):
    plt.xticks(index + bar_width, ('Hog', 'ResNet18', 'Raw'), fontsize=14)
    x_label = 'Features Extraction Method'

elif expNumber == 8:
    plt.xticks(index + bar_width / 2, ('JPEG_Compression', 'Mild_Gaussian', 'Strong_Gaussian',
                                       'Motion_Blur', 'Clean', 'Full'), fontsize=14)
    x_label = 'Test Set'
    first_label = 'Hog'
    second_label = 'ResNet18'

if expNumber not in (3, 4, 6, 7, 10):
    rects1 = plt.bar(index, all_mse1, bar_width,
                     alpha=opacity,
                     color='r',
                     label=first_label)
    autolabel(rects1, all_mse1)

rects2 = plt.bar(index + bar_width, all_mse2, bar_width,
                 alpha=opacity,
                 color='b',
                 label=second_label)
autolabel(rects2, all_mse2)

major_ticks = np.arange(0, 1.5, 0.1)
minor_ticks = np.arange(0, 1.1, 0.1)

ax.set_yticks(major_ticks)
ax.set_yticks(minor_ticks, minor=True)

title = 'MSE - ' + part_of_title
plt.title(title, fontsize=16)
plt.xlabel(x_label, fontsize=18)
plt.ylabel('MSE', fontsize=18)
ax.legend(loc='upper right', bbox_to_anchor=(1, 1), fontsize=14)
box = ax.get_position()
ax.set_position([box.x0, box.y0, box.width * 0.8, box.height])


plt.show()
